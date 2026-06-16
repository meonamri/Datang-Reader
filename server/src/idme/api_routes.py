"""
IDME API Routes (Flask Blueprint)

Provides REST API endpoints and the settings Web UI for the IDME module.

Endpoints:
  GET  /idme/status              - Module status
  GET  /idme/settings            - Teacher management Web UI
  POST /idme/teachers            - Add a teacher
  DELETE /idme/teachers/<id>     - Delete a teacher
  POST /idme/teachers/<id>/test  - Test IDME login credentials
  GET  /idme/scans               - Today's scan summary
  POST /idme/submit              - Manual IDME submission trigger
  GET  /idme/roster              - Student roster summary
  POST /idme/roster/import       - Import roster from Excel (path on server)
  POST /idme/roster/upload       - Upload + import roster Excel (optional replace)
  GET  /idme/roster/template     - Download a blank roster Excel template
  GET  /idme/submissions         - Submission history
"""

import logging
import asyncio
import os
import tempfile
from io import BytesIO
from datetime import date
from pathlib import Path
from flask import Blueprint, request, jsonify, render_template_string, send_file

from .idme_config import IDMEConfig

logger = logging.getLogger(__name__)

# Create blueprint
idme_bp = Blueprint('idme', __name__, url_prefix='/idme')

# Module-level references (set during initialization)
_orchestrator = None
_teacher_manager = None
_roster_manager = None
_scan_tracker = None
_absence_detector = None
_scheduler = None


def init_idme_module(service_manager=None):
    """
    Initialize the IDME module and return the orchestrator.

    Called from http_server.py during Flask app setup.
    Sets up all IDME components and optionally starts the scheduler.

    Args:
        service_manager: Existing Datang ServiceManager (for scan hook).

    Returns:
        IDMEOrchestrator instance.
    """
    global _orchestrator, _teacher_manager, _roster_manager
    global _scan_tracker, _absence_detector, _scheduler

    if not IDMEConfig.ENABLED:
        logger.info("IDME module disabled (IDME_ENABLED=false)")
        return None

    # Validate config
    is_valid, errors = IDMEConfig.validate()
    if not is_valid:
        for err in errors:
            logger.error(f"IDME config error: {err}")
        return None

    from .orchestrator import IDMEOrchestrator
    from .scheduler import IDMEScheduler

    # Initialize orchestrator (creates all sub-components)
    _orchestrator = IDMEOrchestrator(IDMEConfig.DATABASE_PATH)
    _teacher_manager = _orchestrator.teacher_manager
    _roster_manager = _orchestrator.roster_manager
    _scan_tracker = _orchestrator.scan_tracker
    _absence_detector = _orchestrator.absence_detector

    # Hook scan tracker into existing Datang ServiceManager
    if service_manager:
        service_manager.scan_tracker = _scan_tracker
        logger.info("IDME scan tracker hooked into ServiceManager")

    # Start scheduler
    _scheduler = IDMEScheduler(_orchestrator, IDMEConfig.CUTOFF_TIME)
    _scheduler.start()

    logger.info("IDME module initialized successfully")
    return _orchestrator


# ============================================================
# Status endpoint
# ============================================================

@idme_bp.route('/status', methods=['GET'])
def idme_status():
    """Get IDME module status."""
    if not _orchestrator:
        return jsonify({'enabled': False, 'message': 'IDME module not initialized'}), 200

    teachers = _teacher_manager.get_all_teachers() if _teacher_manager else []
    scans_today = _scan_tracker.get_all_scans_today() if _scan_tracker else {}

    status = {
        'enabled': True,
        'config': IDMEConfig.to_dict(),
        'teachers': len(teachers),
        'roster': {
            'total_students': _roster_manager.get_total_students() if _roster_manager else 0,
            'total_classes': _roster_manager.get_total_classes() if _roster_manager else 0,
        },
        'scans_today': scans_today,
        'scheduler': _scheduler.get_status() if _scheduler else None,
    }

    return jsonify(status), 200


# ============================================================
# Settings Web UI
# ============================================================

def _build_overview():
    """
    Aggregate the operational state the settings UI needs in one shot:
    safety mode (enabled / draft-vs-locked), next scheduled fire, and a
    per-class alignment + today preview (roster vs scans vs absent).

    Pure read; tolerant of an uninitialised module (every dependency is
    None-guarded so this is safe to call when IDME is disabled). Shared by
    both ``settings_page`` (server-render, survives a JS failure) and the
    ``/overview`` endpoint (client refresh).
    """
    teachers = _teacher_manager.get_all_teachers(include_disabled=True) if _teacher_manager else []
    classes = _roster_manager.get_all_classes() if _roster_manager else []
    scans = _scan_tracker.get_all_scans_today() if _scan_tracker else {}
    scheduler = _scheduler.get_status() if _scheduler else None

    roster_class_names = {c['class_name'] for c in classes}

    # Per-class alignment + today preview. A class is "onboarded" (will be
    # submitted to MOEIS at cutoff) only when it has an *enabled* teacher.
    class_rows = []
    for c in classes:
        cn = c['class_name']
        teacher = next((t for t in teachers if t['class_name'] == cn), None)
        roster_n = c.get('student_count', 0)
        scanned = scans.get(cn, 0)
        class_rows.append({
            'class_name': cn,
            'roster': roster_n,
            'scanned': scanned,
            'absent': max(roster_n - scanned, 0),
            'teacher_name': teacher['name'] if teacher else None,
            'teacher_enabled': bool(teacher['enabled']) if teacher else False,
            'onboarded': bool(teacher and teacher['enabled']),
        })

    # Teachers whose class_name matches no roster class — the silent-misfire
    # case (CLAUDE.md: the class string must match in three places).
    orphan_teachers = [
        {'id': t['id'], 'name': t['name'], 'class_name': t['class_name']}
        for t in teachers if t['class_name'] not in roster_class_names
    ]

    # Scan sections (the third leg of the three-way match) that hit no roster
    # class — taps landing under a section string nothing else knows about.
    orphan_sections = sorted(s for s in scans if s not in roster_class_names)

    return {
        'config': IDMEConfig.to_dict(),
        'scheduler': scheduler,
        'classes': class_rows,
        'orphan_teachers': orphan_teachers,
        'orphan_sections': orphan_sections,
    }


@idme_bp.route('/overview', methods=['GET'])
def overview():
    """Operational overview (safety mode + per-class alignment/today preview)."""
    return jsonify(_build_overview()), 200


@idme_bp.route('/settings', methods=['GET'])
def settings_page():
    """Render the teacher management Web UI."""
    # Read template
    template_path = Path(__file__).parent / 'templates' / 'settings.html'
    template_content = template_path.read_text()

    teachers = _teacher_manager.get_all_teachers(include_disabled=True) if _teacher_manager else []
    total_students = _roster_manager.get_total_students() if _roster_manager else 0
    total_classes = _roster_manager.get_total_classes() if _roster_manager else 0

    return render_template_string(
        template_content,
        teachers=teachers,
        total_students=total_students,
        total_classes=total_classes,
        cutoff_time=IDMEConfig.CUTOFF_TIME,
        overview=_build_overview(),
    )


# ============================================================
# Teacher management
# ============================================================

@idme_bp.route('/teachers', methods=['POST'])
def add_teacher():
    """Add a new teacher."""
    if not _teacher_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    data = request.get_json()
    if not data:
        return jsonify({'error': 'JSON body required'}), 400

    required = ['name', 'ic_number', 'password', 'class_name']
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({'error': f'Missing fields: {", ".join(missing)}'}), 400

    try:
        teacher = _teacher_manager.add_teacher(
            name=data['name'],
            ic_number=data['ic_number'],
            password=data['password'],
            class_name=data['class_name'],
        )
        return jsonify(teacher), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@idme_bp.route('/teachers/<int:teacher_id>', methods=['PUT'])
def update_teacher(teacher_id):
    """Update a teacher."""
    if not _teacher_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    data = request.get_json()
    if not data:
        return jsonify({'error': 'JSON body required'}), 400

    try:
        teacher = _teacher_manager.update_teacher(
            teacher_id,
            name=data.get('name'),
            ic_number=data.get('ic_number'),
            password=data.get('password'),
            class_name=data.get('class_name'),
            enabled=data.get('enabled'),
        )
        return jsonify(teacher), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@idme_bp.route('/teachers/<int:teacher_id>', methods=['DELETE'])
def delete_teacher(teacher_id):
    """Delete a teacher."""
    if not _teacher_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    try:
        deleted = _teacher_manager.delete_teacher(teacher_id)
        if deleted:
            return jsonify({'success': True}), 200
        else:
            return jsonify({'error': 'Teacher not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@idme_bp.route('/teachers/<int:teacher_id>/test', methods=['POST'])
def test_teacher(teacher_id):
    """Test IDME login with teacher's credentials."""
    if not _teacher_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    try:
        from .login_engine import IDMELoginEngine
        creds = _teacher_manager.get_teacher_credentials(teacher_id)

        async def _test():
            engine = IDMELoginEngine(
                ic_number=creds['ic_number'],
                password=creds['password'],
                headless=True,
                debug=False,
            )
            try:
                result = await engine.login_and_navigate()
                return {
                    'success': True,
                    'duration': f"{result.get('duration', 0):.1f}s",
                    'cookies': len(result.get('cookies', [])),
                    'csrf_token': bool(result.get('csrf_token')),
                }
            finally:
                await engine.close()

        result = asyncio.run(_test())
        return jsonify(result), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 200


# ============================================================
# Scans
# ============================================================

@idme_bp.route('/scans', methods=['GET'])
def get_scans():
    """Get today's scan summary."""
    if not _absence_detector:
        return jsonify({'error': 'IDME module not initialized'}), 503

    class_name = request.args.get('class')
    scan_date = request.args.get('date', date.today().isoformat())

    if class_name:
        summary = _absence_detector.get_attendance_summary(class_name, scan_date)
        return jsonify(summary), 200
    else:
        # All classes summary
        scans = _scan_tracker.get_all_scans_today()
        classes = _roster_manager.get_all_classes() if _roster_manager else []

        result = {
            'date': scan_date,
            'classes': {},
        }
        for cls in classes:
            cls_name = cls['class_name']
            result['classes'][cls_name] = {
                'roster_count': cls['student_count'],
                'scanned_count': scans.get(cls_name, 0),
            }

        return jsonify(result), 200


# ============================================================
# Submit
# ============================================================

@idme_bp.route('/submit', methods=['POST'])
def submit_to_idme():
    """Manually trigger IDME submission for a class."""
    if not _orchestrator:
        return jsonify({'error': 'IDME module not initialized'}), 503

    data = request.get_json() or {}
    class_name = data.get('class_name')
    submission_date = data.get('date', date.today().isoformat())
    # Manual triggers default to a re-editable DRAFT (confirm=False) — the safer
    # path. Pass {"confirm": true} to hard-confirm (TELAH DISAHKAN). The daily
    # scheduler submits with IDMEConfig.SCHEDULER_CONFIRM (drafts by default).
    confirm = bool(data.get('confirm', False))

    if class_name:
        # Submit for specific class
        teacher = _teacher_manager.get_teacher_for_class(class_name)
        if not teacher:
            return jsonify({'error': f'No teacher configured for class {class_name}'}), 400

        try:
            result = _orchestrator.submit_class(
                teacher_id=teacher['id'],
                class_name=class_name,
                submission_date=submission_date,
                confirm=confirm,
            )
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'error': str(e), 'status': 'failed'}), 500
    else:
        # Submit all configured classes
        try:
            results = _orchestrator.submit_all_classes(submission_date, confirm=confirm)
            return jsonify({'results': results}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500


# ============================================================
# Roster
# ============================================================

@idme_bp.route('/roster', methods=['GET'])
def get_roster():
    """Get student roster summary."""
    if not _roster_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    class_name = request.args.get('class')

    if class_name:
        students = _roster_manager.get_class_roster(class_name)
        return jsonify({
            'class_name': class_name,
            'count': len(students),
            'students': students,
        }), 200
    else:
        classes = _roster_manager.get_all_classes()
        return jsonify({
            'total_students': _roster_manager.get_total_students(),
            'total_classes': len(classes),
            'classes': classes,
        }), 200


@idme_bp.route('/roster/coverage', methods=['GET'])
def roster_coverage():
    """
    Tag-mapping coverage for the settings UI (read-only aggregate; no PII beyond
    names already shown in the roster view). Reports mapped/total per class and
    school-wide, the still-unmapped students, and pending unmatched/ambiguous
    scan counts. See IDENTITY_RESOLUTION_DESIGN.md §5.4.
    """
    if not _roster_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    coverage = _roster_manager.get_tag_coverage()
    coverage['unmatched'] = (
        _scan_tracker.get_unmatched_summary() if _scan_tracker
        else {'no_match': 0, 'ambiguous': 0, 'total': 0}
    )
    return jsonify(coverage), 200


@idme_bp.route('/roster/init', methods=['POST'])
def init_roster():
    """
    Initialise (seed/refresh) the identity registry for a class from the MOEIS
    portal. READ-ONLY against the portal — logs in, reads the student table, and
    upserts into the local registry. Learned RFID tags are preserved.

    JSON body: {"class_name": "5 UKM"}  (a teacher must be configured for it).
    """
    if not _orchestrator or not _teacher_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    data = request.get_json() or {}
    class_name = data.get('class_name')
    if not class_name:
        return jsonify({'error': 'class_name is required'}), 400

    teacher = _teacher_manager.get_teacher_for_class(class_name)
    if not teacher:
        return jsonify({'error': f'No teacher configured for class {class_name}'}), 400

    try:
        result = _orchestrator.init_roster_from_portal(
            teacher_id=teacher['id'],
            class_name=class_name,
        )
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e), 'status': 'failed'}), 500


@idme_bp.route('/roster/import', methods=['POST'])
def import_roster():
    """Import student roster from Excel file."""
    if not _roster_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    data = request.get_json() or {}
    excel_path = data.get('excel_path', '')

    if not excel_path:
        return jsonify({'error': 'excel_path is required'}), 400

    try:
        result = _roster_manager.import_from_excel(excel_path)
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# Roster Excel template columns (matched case-insensitively on import)
ROSTER_TEMPLATE_COLUMNS = ['Name', 'Class', 'IC', 'Tag', 'ID']
MAX_ROSTER_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB


@idme_bp.route('/roster/upload', methods=['POST'])
def upload_roster():
    """
    Upload an Excel file from the browser and import it.

    Multipart form fields:
      - file:    the .xlsx/.xls file
      - replace: 'true' to clear the existing roster before importing
                 (avoids duplicate students on re-import)
    """
    if not _roster_manager:
        return jsonify({'error': 'IDME module not initialized'}), 503

    if request.content_length and request.content_length > MAX_ROSTER_UPLOAD_BYTES:
        return jsonify({'error': 'File too large (max 10 MB)'}), 413

    file = request.files.get('file')
    if file is None or not file.filename:
        return jsonify({'error': 'No file uploaded (field name must be "file")'}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File must be an Excel .xlsx or .xls'}), 400

    replace = request.form.get('replace', 'false').lower() == 'true'

    # Save to a temp file so the existing path-based importer can read it
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()
    try:
        file.save(tmp_path)

        cleared = _roster_manager.clear_roster() if replace else 0
        result = _roster_manager.import_from_excel(tmp_path)
        result['replaced'] = replace
        result['cleared'] = cleared
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@idme_bp.route('/roster/template', methods=['GET'])
def roster_template():
    """Generate and download a blank roster Excel template with sample rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    ws.append(ROSTER_TEMPLATE_COLUMNS)
    # Example rows (clearly marked — admin should delete before importing)
    ws.append(['(EXAMPLE) AHMAD BIN ALI', '5 UKM', '120101010101', '1234567890', 'S001'])
    ws.append(['(EXAMPLE) SITI BINTI ABU', '5 UKM', '120202020202', '1234567891', 'S002'])

    # Style: bold header, sensible column widths
    for col_idx in range(1, len(ROSTER_TEMPLATE_COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    for col_idx, width in enumerate([30, 12, 16, 16, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='roster_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ============================================================
# Submissions
# ============================================================

@idme_bp.route('/submissions', methods=['GET'])
def get_submissions():
    """Get submission history."""
    if not _orchestrator:
        return jsonify({'error': 'IDME module not initialized'}), 503

    class_name = request.args.get('class')
    limit = request.args.get('limit', 50, type=int)

    history = _orchestrator.get_submission_history(class_name, limit)
    return jsonify({'submissions': history}), 200
